"""Delimited text and spreadsheet viewers."""

from __future__ import annotations

import csv
from pathlib import Path

from PyQt6.QtWidgets import QComboBox, QLabel, QTableWidget, QVBoxLayout

from vibe_viewer.viewers.base import BaseViewer, ViewerError
from vibe_viewer.viewers.helpers import fill_table, read_text_safely


class DelimitedTableViewer(BaseViewer):
    name = "Delimited tables"
    category = "Tables"
    priority = 96
    extensions = (".csv", ".tsv", ".tab")

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self.info = QLabel()
        self.table = QTableWidget()
        layout = QVBoxLayout(self)
        layout.setContentsMargins(6, 6, 6, 6)
        layout.addWidget(self.info)
        layout.addWidget(self.table)

    def load_file(self, path: Path) -> None:
        text, encoding, text_truncated = read_text_safely(path, limit=16 * 1024 * 1024)
        try:
            sample = text[:8192]
            delimiter = "\t" if path.suffix.lower() in {".tsv", ".tab"} else None
            if delimiter:
                dialect = csv.excel_tab
            else:
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except csv.Error:
                    dialect = csv.excel
            rows = csv.reader(text.splitlines(), dialect=dialect)
            row_count, column_count, rows_truncated = fill_table(self.table, rows)
        except csv.Error as exc:
            raise ViewerError(f"Не удалось разобрать таблицу: {exc}") from exc
        notes = []
        if text_truncated:
            notes.append("файл обрезан до 16 МБ")
        if rows_truncated:
            notes.append("показаны первые 2000 строк")
        suffix = f" • {', '.join(notes)}" if notes else ""
        self.info.setText(
            f"{path.name} • {encoding} • {row_count}×{column_count}{suffix}"
        )


class SpreadsheetViewer(BaseViewer):
    name = "Spreadsheets"
    category = "Office and spreadsheets"
    priority = 100
    extensions = (
        ".xlsx", ".xlsm", ".xltx", ".xltm", ".xls", ".ods", ".ots",
        ".dbf", ".orc", ".avro", ".xlsb", ".dif", ".slk",
    )

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self._path: Path | None = None
        self._book = None
        self.info = QLabel()
        self.sheet_selector = QComboBox()
        self.sheet_selector.currentIndexChanged.connect(self._show_current_sheet)
        self.table = QTableWidget()
        layout = QVBoxLayout(self)
        layout.setContentsMargins(6, 6, 6, 6)
        layout.addWidget(self.info)
        layout.addWidget(self.sheet_selector)
        layout.addWidget(self.table)

    def load_file(self, path: Path) -> None:
        self.unload()
        self._path = path
        suffix = path.suffix.lower()
        try:
            if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
                import openpyxl

                self._book = openpyxl.load_workbook(
                    path, read_only=True, data_only=True, keep_links=False
                )
                names = list(self._book.sheetnames)
            elif suffix == ".xls":
                import xlrd

                self._book = xlrd.open_workbook(path, on_demand=True)
                names = self._book.sheet_names()
            elif suffix in {".ods", ".ots"}:
                from odf.opendocument import load

                self._book = load(str(path))
                from odf.table import Table

                names = [str(table.getAttribute("name") or f"Лист {index + 1}") for index, table in enumerate(self._book.getElementsByType(Table))]
            else:
                names = ["Данные"]
        except Exception as exc:
            raise ViewerError(f"Не удалось открыть электронную таблицу: {exc}") from exc

        self.sheet_selector.blockSignals(True)
        self.sheet_selector.clear()
        self.sheet_selector.addItems(names)
        self.sheet_selector.blockSignals(False)
        if names:
            self._show_current_sheet(0)
        else:
            self.info.setText(f"{path.name} • нет листов")

    def _show_current_sheet(self, index: int) -> None:
        if index < 0 or self._path is None:
            return
        suffix = self._path.suffix.lower()
        try:
            if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
                if self._book is None:
                    return
                sheet = self._book[self._book.sheetnames[index]]
                rows = sheet.iter_rows(values_only=True)
            elif suffix == ".xls":
                if self._book is None:
                    return
                sheet = self._book.sheet_by_index(index)
                rows = (sheet.row_values(row) for row in range(sheet.nrows))
            elif suffix in {".ods", ".ots"}:
                if self._book is None:
                    return
                rows = self._odf_rows(index)
            else:
                rows, headers = self._extended_rows(suffix)
                row_count, column_count, truncated = fill_table(self.table, rows, headers=headers)
                note = " • показаны первые 2000 строк" if truncated else ""
                self.info.setText(f"{self._path.name} • {row_count} строк • {column_count} столбцов{note}")
                return
            row_count, column_count, truncated = fill_table(self.table, rows)
            note = " • показаны первые 2000 строк" if truncated else ""
            self.info.setText(
                f"{self._path.name} • {row_count} строк • {column_count} столбцов{note}"
            )
        except Exception as exc:
            raise ViewerError(f"Не удалось показать лист: {exc}") from exc

    def _extended_rows(self, suffix: str):
        if suffix == ".dbf":
            from dbfread import DBF

            table = DBF(str(self._path), load=False, char_decode_errors="replace")
            return ([record.get(name) for name in table.field_names] for record in table), table.field_names
        if suffix == ".orc":
            import pyarrow.orc as orc

            table = orc.ORCFile(self._path).read()
            return table.to_pylist_rows() if hasattr(table, "to_pylist_rows") else (tuple(row.values()) for row in table.to_pylist()), table.column_names
        if suffix == ".avro":
            from fastavro import reader

            with self._path.open("rb") as stream:
                records = reader(stream)
                fields = list(records.writer_schema.get("fields", []))
                names = [field["name"] for field in fields]
                rows = [[record.get(name) for name in names] for record in records]
            return rows, names
        if suffix == ".xlsb":
            from pyxlsb import open_workbook

            with open_workbook(str(self._path)) as book:
                with book.get_sheet(1) as sheet:
                    rows = [[cell.v for cell in row] for row in sheet.rows()]
            return rows, None
        text, _, _ = read_text_safely(self._path, limit=16 * 1024 * 1024)
        if suffix == ".dif":
            return self._dif_rows(text), None
        return self._sylk_rows(text), None

    @staticmethod
    def _dif_rows(text: str) -> list[list[str]]:
        """Decode the DATA section of a Data Interchange Format table."""
        lines = text.splitlines()
        upper_lines = [line.strip().upper() for line in lines]
        try:
            data_index = upper_lines.index("DATA")
        except ValueError as exc:
            raise ViewerError("В DIF отсутствует секция DATA") from exc

        column_count = 0
        try:
            vectors_index = upper_lines.index("VECTORS")
            column_count = int(lines[vectors_index + 1].split(",", 1)[1])
        except (ValueError, IndexError):
            pass

        rows: list[list[str]] = []
        current: list[str] = []
        index = data_index + 3
        while index + 1 < len(lines):
            descriptor = lines[index].strip()
            value_line = lines[index + 1].strip()
            index += 2
            try:
                value_type_text, numeric_value = descriptor.split(",", 1)
                value_type = int(value_type_text)
            except ValueError:
                continue

            marker = value_line.strip().strip('"').upper()
            if value_type == -1:
                if marker == "BOT":
                    if current:
                        rows.append(current)
                    current = []
                elif marker == "EOD":
                    if current:
                        rows.append(current)
                    break
                continue

            if value_type == 1:
                try:
                    value = next(csv.reader([value_line]))[0]
                except (csv.Error, StopIteration):
                    value = value_line.strip('"')
            else:
                value = numeric_value if marker in {"V", "NA", "ERROR"} else value_line.strip('"')
            current.append(value)
            if column_count > 0 and len(current) >= column_count:
                rows.append(current)
                current = []
        if current:
            rows.append(current)
        return rows

    @staticmethod
    def _sylk_rows(text: str) -> list[list[str]]:
        """Decode SYLK C records into a rectangular table."""
        cells: dict[tuple[int, int], str] = {}
        current_x = 1
        current_y = 1
        for line in text.splitlines():
            if not line.startswith("C;"):
                continue
            fields = line.split(";")[1:]
            value: str | None = None
            for field in fields:
                if field.startswith("X") and field[1:].isdigit():
                    current_x = int(field[1:])
                elif field.startswith("Y") and field[1:].isdigit():
                    current_y = int(field[1:])
                elif field.startswith("K"):
                    raw = field[1:]
                    value = raw[1:-1].replace('""', '"') if len(raw) >= 2 and raw.startswith('"') and raw.endswith('"') else raw
            if value is not None and current_x > 0 and current_y > 0:
                cells[(current_y, current_x)] = value
        if not cells:
            raise ViewerError("В SYLK не найдены ячейки")
        max_row = min(max(row for row, _ in cells), 2000)
        max_column = min(max(column for _, column in cells), 200)
        return [
            [cells.get((row, column), "") for column in range(1, max_column + 1)]
            for row in range(1, max_row + 1)
        ]

    def _odf_rows(self, index: int):
        from odf import teletype
        from odf.table import Table, TableCell, TableRow

        table = self._book.getElementsByType(Table)[index]
        for row in table.getElementsByType(TableRow):
            values: list[str] = []
            repeat_rows = min(int(row.getAttribute("numberrowsrepeated") or 1), 100)
            for cell in row.getElementsByType(TableCell):
                repeat_columns = min(int(cell.getAttribute("numbercolumnsrepeated") or 1), 200)
                value = teletype.extractText(cell)
                values.extend([value] * repeat_columns)
            for _ in range(repeat_rows):
                yield values

    def unload(self) -> None:
        if self._book is not None and hasattr(self._book, "close"):
            self._book.close()
        self._book = None
        self._path = None
        self.sheet_selector.clear()
        self.table.clear()
